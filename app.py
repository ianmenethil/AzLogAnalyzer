# app.py  # pylint disable=too-many-lines # pylint: disable=C0302
import logging
import json
import os
from openpyxl import load_workbook
from typing import Any, Tuple, List, Dict, cast
import sys
from pathlib import Path
import pandas as pd
import httpx
from configurator import setup_logging, console, ExcelConfigurations, RegexConfigurations, AppConfigurations, LF

setup_logging()
from compareHeaders import compare_and_write_output  # pylint: disable=wrong-import-position  # pylint: disable=E402
from timeManager import TimeManager  # pylint: disable=wrong-import-position  # pylint: disable=E402
from configLoader import ConfigLoader  # pylint: disable=wrong-import-position  # pylint: disable=E402
from userInputManager import UserInputHandler  # pylint: disable=wrong-import-position  # pylint: disable=E402
from excelManager import ExcelManager  # pylint: disable=wrong-import-position  # pylint: disable=E402
# from backupManager import BackupManager  # pylint: disable=wrong-import-position  # pylint: disable=E402
from fileManager import FileManager  # pylint: disable=wrong-import-position  # pylint: disable=E402
from regexManager import RegexManager  # pylint: disable=wrong-import-position  # pylint: disable=E402
from logManager import LogManager  # pylint: disable=wrong-import-position # pylint: disable=E402

logger = logging.getLogger(__name__)
get_current_time_info = TimeManager.get_current_time_info
convert_utc_to_sydney = TimeManager.convert_utc_to_sydney

OUTPUT_FOLDER: str = 'AZLOGS'
LOG_FOLDER: str = f'{OUTPUT_FOLDER}/LOG'
CL_TimeGenerated: str = 'TimeGenerated'
CL_LocalTime: str = 'LocalTime'
DELALL_FEXT: Tuple[str, str, str, str] = ('.csv', '.xlsx', '.log', '.txt')


class APIManager():

    @staticmethod
    def get_logs_from_azure_analytics(query, headers, zen_endpoint) -> Any | dict:  # ! API call to get logs
        """The `get_logs_from_azure_analytics` function makes an API call to retrieve logs from Azure Analytics
        and returns the response as a JSON object."""
        try:
            logger.info('Calling Azure with following query')
            console.print(query, style='blink', justify='left')
            LF()
            headers['Content-Type'] = 'application/json'
            response = httpx.post(url=zen_endpoint, data=query, headers=headers, timeout=60)
            response.raise_for_status()
            logger.info(f"Response Status Code: {response.status_code}")
            return response.json()
        except httpx.HTTPError as e:
            logger.error(f"HTTP error occurred: {e}", exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)
        except Exception as e:
            logger.error(f"An unexpected error occurred: {e}", exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)
        return {}

    @staticmethod
    def save_new_token(token_url, client_id, client_secret, resource_scope, token_file):
        """The `save_new_token` function sends a request to obtain a new token using client credentials, saves
        the token response to a file, and returns the token response."""
        try:
            headers = {'Content-Type': 'application/x-www-form-urlencoded'}
            payload = {
                'grant_type': 'client_credentials',
                'client_id': client_id,
                'client_secret': client_secret,
                'resource': resource_scope,
            }
            response = httpx.post(token_url, headers=headers, data=payload)
            response.raise_for_status()
            token_response = response.json()
            FileManager.save_json(token_response, token_file)
            return token_response
        except httpx.HTTPError as e:
            logger.error(f"HTTP error occurred: {e}", exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)
        except Exception as e:
            logger.error(f"An error occurred: {str(object=e)}")
        return None, None

    @staticmethod
    def fetch_and_save_api_data(token_url: str = '',
                                client_id: str = '',
                                client_secret: str = '',
                                resource_scope: str = '',
                                token_file_path: str = '',
                                json_file_path: str = '',
                                endpoint: str = ''):
        """The function fetches and saves API data using a token and query, and handles errors."""
        query_name, query_content = UserInputHandler.get_query_input()
        KQL = json.dumps({"query": query_content})
        console.print(KQL)
        logger.info(f'Active [yellow]Query Name:[/yellow] [red]{query_name}[/red]')
        LF()
        console.print('([red]Formatted[/red] Query):', style="blue blink", justify="left")
        LF()
        console.print(query_content, style="white", justify="left")
        LF()
        console.print('IS THE QUERY CORRECT?', style="green bold", justify="left")
        LF()
        correct_query = input(f'Query name: {query_name} Enter to continue, or type anything to exit...')
        LF()
        if correct_query == '':
            logger.info(f'Proceed with: {query_name}')
        else:
            logger.info('Incorrect query, exiting...')
            sys.exit()
        token_info = APIManager.save_new_token(token_url, client_id, client_secret, resource_scope, token_file_path)
        if not token_info:
            logger.warning("Failed to refresh access token.")
            return 'Error in token', query_name
        try:
            token_key = token_info["access_token"]  # type: ignore
            headers = {'Authorization': f'Bearer {token_key}'}

            if response := APIManager.get_logs_from_azure_analytics(KQL, headers, endpoint):
                FileManager.save_json(response, json_file_path)
                return response, query_name
            logger.error("Could not fetch data from API. Exiting.")
            sys.exit()
        except Exception as e:
            logger.error(f"e in fetch_and_save_api_data: {e}", exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)
            logger.error("Could not fetch data from API. Exiting.")
            sys.exit()


class FileHandler():

    @staticmethod
    def add_DF_to_CSVFILE_dropNA(json_data, filename: str, log_file: str):
        # sourcery skip: pandas-avoid-inplace
        """Converts tables in JSON response into pandas DataFrames, drops any blank columns, and saves the resulting DataFrames as CSV files."""
        try:
            if 'tables' in json_data:
                for table in json_data['tables']:
                    df = pd.DataFrame(table['rows'], columns=[col['name'] for col in table['columns']])
                    dataframe_row_length = len(df.index)
                    dataframe_column_length = len(df.columns)
                    dataframe_column_names_before_drop = set(df.columns)
                    logger.info(f'Before dropna, DF Rows: {dataframe_row_length}, DF Columns: {dataframe_column_length}')
                    LogManager.save_removed_rows_to_raw_logs(df, log_file)
                    with open(log_file, 'a', encoding='utf-8') as f:
                        f.write('Before dropna,Column name,Number of rows\n')
                        for c in df.columns:
                            f.write(f'Before dropna,{c},{dataframe_row_length}\n')

                    df.dropna(axis=1, how='all', inplace=True)
                    dataframe_column_names_after_drop = set(df.columns)

                    # Logging and writing after dropna
                    logger.info(f'After dropna - DF Rows: {len(df.index)}, DF Columns: {len(df.columns)}')
                    with open(log_file, 'a', encoding='utf-8') as f:
                        f.write('\nAfter dropna,Column name,Number of rows\n')
                        for c in df.columns:
                            f.write(f'After dropna,{c},{len(df.index)}\n')
                    # with open(log_file, 'a', encoding='utf-8') as f:
                    #     f.write('\nAfter dropna - Column name,Number of rows\n')
                    #     for c in df.columns:
                    #         f.write(f'{c},{len(df.index)}\n')

                    # if (columns_dropped := dataframe_column_names_before_drop - dataframe_column_names_after_drop):
                    #     with open(log_file, 'a', encoding='utf-8') as f:
                    #         f.write('\nDropped columns due to all NaN\n')
                    #         for c in columns_dropped:
                    #             f.write(f'{c}\n')
                    if (columns_dropped := dataframe_column_names_before_drop - dataframe_column_names_after_drop):
                        with open(log_file, 'a', encoding='utf-8') as f:
                            f.write('\nDropped columns due to all NaN,Column name\n')
                            for c in columns_dropped:
                                f.write(f'Dropped columns due to all NaN,{c}\n')
                    # Saving the DataFrame to CSV after dropping NaN columns
                    df.to_csv(filename, index=False, encoding='utf-8')
                    logger.info(f'Saved {filename}')
                    return True
            else:
                logger.warning("No tables found in response.")
                return None
        except KeyError as e:
            logger.error(f"Key error in processing JSON data: {e}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return None

    @staticmethod
    def add_LocalTime_col(input_file, source_col, dest_col, output_file) -> None:
        try:
            df = pd.read_csv(input_file, low_memory=False)
            logger.info(f"source_col: {source_col} | dest_col: {dest_col}")
            new_column_data = df[source_col].apply(convert_utc_to_sydney)
            if source_col not in df.columns:
                logger.warning(f"Column {source_col} not found in DF.")
            # ? Add LocalTime column
            else:
                logger.info(f"Column {source_col} found in DF. Adding {dest_col} column to DF.")
                df.insert(0, dest_col, new_column_data)
            df.to_csv(output_file, index=False)
            logger.info(f"DataFrame saved to CSV with new order: {output_file}")
        except Exception as e:
            logger.error(f'Error in processing data: {e}', exc_info=True, stack_info=True)


class DataProcessor():

    @staticmethod
    def process_data(JSON_FILE, CSV_FILE, LOGS_FILE, EXCEL_FILE) -> Any:
        try:
            JSONDATA = FileManager.read_json(filename=JSON_FILE)
            if 'tables' in JSONDATA:
                saveData = FileHandler.add_DF_to_CSVFILE_dropNA(json_data=JSONDATA, filename=CSV_FILE, log_file=LOGS_FILE)
                if saveData:
                    logger.info(f'CSV file created: {CSV_FILE}')
                else:
                    logger.error(f'CSV file not created: {CSV_FILE}. Exiting...')
                    sys.exit()
                if saveData:
                    FileHandler.add_LocalTime_col(
                        input_file=CSV_FILE,
                        source_col=CL_TimeGenerated,
                        dest_col=CL_LocalTime,
                        output_file=CSV_FILE,
                    )
                    logger.info(f'LT column added to {CSV_FILE}')
                    ExcelManager.excelCreate(input_csv_file=CSV_FILE, output_excel_file=EXCEL_FILE)
                    logger.info(f'excelCreate: Input {CSV_FILE} | Output {EXCEL_FILE}')
                    return EXCEL_FILE
        except Exception as e:
            logger.error(f'E in process_local_data: {e}', exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)

    @staticmethod
    def reorder_sheets(excel_file: str, focused_cols_set: list) -> None:
        final_sheet_dataframe = pd.read_excel(excel_file, sheet_name='Sheet1')
        columns_to_include = [col for col in focused_cols_set if col in final_sheet_dataframe.columns]
        focused_df = final_sheet_dataframe[columns_to_include]
        wb = load_workbook(excel_file)

        if 'AllData' not in wb.sheetnames:
            all_data_sheet = wb.create_sheet('AllData')
            sheet1 = wb['Sheet1']
            for row in sheet1.iter_rows(values_only=True):
                all_data_sheet.append(row)
            del wb['Sheet1']

        wb.save(excel_file)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
            focused_df.to_excel(writer, sheet_name='FocusedColumns', index=False)

        wb = load_workbook(excel_file)
        sheets = wb.sheetnames
        focused_idx = sheets.index('FocusedColumns')
        sheets = [sheets.pop(focused_idx)] + sheets  # Move 'FocusedColumns' to the first position
        wb._sheets = [wb[s] for s in sheets]  # type: ignore
        wb.save(excel_file)


def main() -> None:
    """The main function loads configurations, processes data from either a local file or an API, and
    manages backups."""
    excel_config, regex_config, az_config = ConfigLoader.load_configurations()
    excel_config = cast(ExcelConfigurations, excel_config)
    regex_config = cast(RegexConfigurations, regex_config)
    az_config = cast(AppConfigurations, az_config)

    # # ! Columns
    CL_DROPPED: List[str] = excel_config['CL_DROPPED']
    CL_FINAL_ORDER: List[str] = excel_config['CL_FINAL_ORDER']
    MATCH_VALUE_TO_SPECIFIC_COLUMN: Dict[str, List[str]] = regex_config['MATCH_VALUE_TO_SPECIFIC_COLUMN']

    # # ! Configs
    TENANT_ID: str = az_config['TENANT_ID']
    CLIENT_ID: str = az_config['CLIENT_ID']
    CLIENT_SECRET: str = az_config['CLIENT_SECRET']
    ZEN_WORKSPACE_ID: str = az_config['ZEN_WORKSPACE_ID']
    AZDIAG_SCOPE: str = az_config['AZDIAG_SCOPE']
    AZLOG_ENDPOINT: str = az_config['AZLOG_ENDPOINT']
    TOKEN_FILEPATH_STR: str = az_config['TOKEN_FILEPATH_STR']
    JSON_FILE: str = az_config['AZDIAG_JSON_FILEPATH_STR']
    CSV_FILE: str = az_config['AZDIAG_CSV_FILEPATH_STR']
    EXCEL_FILE: Path = Path(az_config['AZDIAG_EXCEL_FILEPATH'])
    EXCEL_FINAL_FILE: Path = Path(az_config['AZDIAG_EXCEL_FINAL_FILEPATH'])
    AZDIAG_EXCEL_COMBINED: Path = Path(az_config['AZDIAG_EXCEL_COMBINED_FILE'])
    TMPEXCEL: Path = Path(az_config['TEMP_EXCEL_FILE'])
    LOG_FILE: str = az_config['Extraction_LogFILE']
    # TOKEN_URL: str = f'https://login.microsoftonline.com/{TENANT_ID}/oauth2/token'
    # AZLOG_ZEN_ENDPOINT: str = f'{AZLOG_ENDPOINT}{ZEN_WORKSPACE_ID}/query'
    IP_DIR = 'IPCheck'
    AZLOG_ZEN_ENDPOINT: str = 'http://127.0.0.1:5000/query'
    TOKEN_URL: str = 'http://127.0.0.1:5000/token'

    if Path(JSON_FILE).is_file():
        new_data_choice = input('Get data?')
        if new_data_choice == '':
            new_data_choice = 'n'
            logger.info(f'Choice: {new_data_choice} - No new data will be fetched.')
        else:
            logger.info(f'Choice: {new_data_choice} - New data will be fetched.')
            Path(JSON_FILE).unlink(missing_ok=True)
    logger.info(f'Fetching data from: {JSON_FILE}')
    if Path(TOKEN_FILEPATH_STR).is_file():
        Path(TOKEN_FILEPATH_STR).unlink(missing_ok=True)
    FileManager.remove_files_from_folder(folder=OUTPUT_FOLDER, file_extension=DELALL_FEXT)
    logger.info(f'Removed files from {OUTPUT_FOLDER} with extension: {DELALL_FEXT}')
    LF()
    if Path(JSON_FILE).is_file():
        DataProcessor.process_data(JSON_FILE, CSV_FILE, LOG_FILE, EXCEL_FILE)
        logger.info(f'Local data source: {JSON_FILE}')
        logger.info(f'Saved to: {EXCEL_FILE}')
    else:
        data, query_name = APIManager.fetch_and_save_api_data(TOKEN_URL, CLIENT_ID, CLIENT_SECRET, AZDIAG_SCOPE, TOKEN_FILEPATH_STR, JSON_FILE, AZLOG_ZEN_ENDPOINT)
        logger.info(f'Query name: {query_name}')
        logger.info(f'API data saved to: {JSON_FILE}')
        if data:
            DataProcessor.process_data(JSON_FILE, CSV_FILE, LOG_FILE, EXCEL_FILE)
            logger.info(f'API data from source: {JSON_FILE}')
            logger.info(f'Saved to: {EXCEL_FILE}')
        else:
            logger.error('Failed to process API data.')

    try:
        logger.info(f'Excel file loaded from: {EXCEL_FILE}')
        data = pd.read_excel(EXCEL_FILE)
    except Exception as e:
        logger.error(f'Error in read_excel exiting: {e}', exc_info=True, stack_info=True)
        sys.exit()

    # ! Filter based on pattern
    try:
        logger.info(f'Filtered based on patterns: {MATCH_VALUE_TO_SPECIFIC_COLUMN}')
        RegexManager.filter_df_based_on_patterns(df=data, config_dict=MATCH_VALUE_TO_SPECIFIC_COLUMN, output_dir=OUTPUT_FOLDER)
    except Exception as e:
        logger.error(f'Error in filter_df_based_on_patterns: {e}', exc_info=True, stack_info=True)
    try:
        EXCEL_FILE_STR = str(EXCEL_FILE)
        EXCEL_FINAL_FILE_STR = str(EXCEL_FINAL_FILE)
        ExcelManager.create_final_excel(input_file=EXCEL_FILE_STR,
                                        output_file=EXCEL_FINAL_FILE_STR,
                                        logfile=LOG_FILE,
                                        columns_to_be_dropped=CL_DROPPED,
                                        final_column_order=CL_FINAL_ORDER)
        logger.info(f'Excel file saved to: {EXCEL_FINAL_FILE}')
    except Exception as e:
        logger.error(f'Error in create_final_excel: {e}', exc_info=True, stack_info=True)

    # try:
    #     df = pd.read_excel(EXCEL_FILE)  # type: ignore
    #     logger.info(f'Excel file loaded: {EXCEL_FILE}')
    try:
        df = pd.read_excel(EXCEL_FINAL_FILE)  # type: ignore
        logger.info(f'Excel file loaded: {EXCEL_FINAL_FILE}')
        column_names = df.columns.tolist()
        df.dropna(axis=1, how='all', inplace=True)
        col_names_after_drop = df.columns.tolist()
        dropped_columns = list(set(column_names) - set(col_names_after_drop))
        logger.info(f'Columns dropped: {dropped_columns}')
        regexdf = RegexManager.filter_df_based_on_patterns(df=df, config_dict=MATCH_VALUE_TO_SPECIFIC_COLUMN, output_dir=OUTPUT_FOLDER)
        temp_path: str = str(TMPEXCEL)
        regexdf.to_excel(temp_path, index=False)
    except Exception as e:
        logger.error(f'read_excel: {e}', exc_info=True, stack_info=True)

    try:
        temp_path = str(TMPEXCEL)
        ExcelManager.create_final_excel(input_file=temp_path,
                                        output_file=str(EXCEL_FINAL_FILE),
                                        logfile=LOG_FILE,
                                        columns_to_be_dropped=CL_DROPPED,
                                        final_column_order=CL_FINAL_ORDER)
    except Exception as e:
        logger.error(f'Error in reading Excel file: {e}', exc_info=True, stack_info=True)
    LF()
    # cs = os.path.splitext(CSV_FILE)
    # ex = os.path.splitext(EXCEL_FINAL_FILE)
    # COMPARE_HEADER_FILE = str(cs) + '_' + str(ex) + '_compare_headers.xlsx'
    # compare_and_write_output(CSV_FILE, EXCEL_FINAL_FILE, COMPARE_HEADER_FILE)
    compare_and_write_output(CSV_FILE, EXCEL_FINAL_FILE)
    LF()
    focused_cols_set = [
        'LocalTime',
        'IPs',
        'userAgent_s',
        'Method_Type',
        'QueryType',
        'QueryParameters',
        'Query',
        'Host',
        'httpStatus_d',
        'serverStatus_s',
        'timeTaken_d',
        'serverResponseLatency_s',
        'WAFMode_s',
        'ruleName_s',
        'clientResponseTime_d',
        'clientResponseLatency_s'  # !
        # 'mPay_OneOffPaymentSuccessful',
        # 'Pay_PublicOneOffPaymentMerchantSearchByCode',
        # 'CYBS_StepUp',
        # 'Pay_InitiatePayment',
        # 'CYBS_getsignedobject',
        # 'Pay_PublicOneOffPaymentConfirmation',
        # 'Authorise',
        # 'GetProcessorInfo',
        # 'mPay_GetTransactionFee',
        # 'CYBS_Direct',
        # 'v3_Authorise',
        # 'Pay_ReviewPayment', #
    ]

    DataProcessor.reorder_sheets(str(EXCEL_FINAL_FILE), focused_cols_set)
    logger.info('Sheet order updated successfully.')

    # !
    source_files = [EXCEL_FINAL_FILE]
    first_file = EXCEL_FILE
    columns = ['clientIP_s', 'clientIp_s']
    new_file = AZDIAG_EXCEL_COMBINED
    chars = '[]"'
    for col in columns:
        unique_data = ExcelManager.read_and_filter_excel(first_file, columns=[col])
        logger.info(f"Unique data collected: {unique_data} from {first_file} for column: {col}")

        cleaned_unique_data = []

        for data in unique_data:
            for char in chars:
                data = data.replace(char, "")
            cleaned_unique_data.append(data)

        unique_data = cleaned_unique_data
        for data in unique_data:
            logger.info(f"Processing IP: {data}")
            ip_file = f"{IP_DIR}/{data}.xlsx"

            if os.path.exists(ip_file):
                source_files.append(Path(ip_file))
        logger.info(f"Source files: {source_files}")
        try:
            ExcelManager.copy_sheets_to_new_file(source_files, new_file)
            if os.path.exists(new_file):
                logger.info(f"New file created: {new_file}")
        except Exception as e:
            logger.error(f'Error in copy_sheets_to_new_file: {e}', exc_info=True, stack_info=True)


if __name__ == '__main__':
    main()
