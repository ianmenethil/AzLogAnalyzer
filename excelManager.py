import logging
from typing import List, Dict, Any
import pandas as pd  # pylint: disable=C0411
from pandas.errors import EmptyDataError
from openpyxl import load_workbook, Workbook
import csv

logger = logging.getLogger(__name__)


class ExcelManager():

    @staticmethod
    def excelCreate(
        input_csv_file,
        output_excel_file,
    ) -> None:
        try:
            df = pd.read_csv(input_csv_file, low_memory=False)
            df = df.dropna(axis=1, how='all')
            if not df.empty:
                with pd.ExcelWriter(output_excel_file, engine='xlsxwriter') as writer:  # pylint: disable=abstract-class-instantiated
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    ExcelManager.format_excel_file(writer, df)
                    logger.info(f'Excel formatted and created: {output_excel_file}')
                    return output_excel_file
            else:
                logger.warning('DF is empty')
        except Exception as e:
            logger.error(f'E in create_excel: {e}', exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)

    @staticmethod
    def format_excel_file(writer, df) -> None:
        """The `format_excel_file` function formats an Excel file by setting the font, font size, and column width for each column in a given DF."""
        try:
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            cell_format = workbook.add_format({'font_name': 'Open Sans', 'font_size': 8})
            worksheet.set_column('A:ZZ', None, cell_format)
            for column in df:
                column_length = max(df[column].astype(str).apply(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                worksheet.set_column(col_idx, col_idx, column_length, cell_format)
        except Exception as e:
            logger.error(f'Error in format_excel_file: {e}', exc_info=True, stack_info=True, extra={'color': 'red'}, stacklevel=2)

    @staticmethod
    def create_final_excel(input_file: str, output_file: str, logfile: str, columns_to_be_dropped: List[str], final_column_order: List[str]):

        def concatenate_values(row: pd.Series, columns: List[str], include_col_names: bool = False) -> str:
            if include_col_names:
                return ', '.join([f'{col}: "{row[col]}"' for col in columns if pd.notna(row[col])])
            return ', '.join([f'"{row[col]}"' for col in columns if pd.notna(row[col])])

        def drop_and_create_columns(df: pd.DataFrame, new_columns: Dict[str, Dict[str, Any]]) -> pd.DataFrame:
            logger.info('[red]########### columns_to_drop started ###########[/red]')
            for col_name, info in new_columns.items():
                existing_cols = [col for col in info['cols'] if col in df.columns]
                missing_cols = [col for col in info['cols'] if col not in df.columns]
                if existing_cols:
                    # ! Create new column by concatenating values from existing columns
                    df[col_name] = df.apply(lambda row, cols=existing_cols: concatenate_values(row, cols, info['include_names']), axis=1)  # pylint: disable=W0640
                    logger.info(f'New column "{col_name}" created with values from columns: {existing_cols} - Total rows: {df.shape[0]}')
                    if col_name != 'Qs':
                        columns_to_drop_due_to_concat.extend(existing_cols)
                    elif col_name != 'Method_Type':
                        columns_to_drop_due_to_concat.extend(existing_cols)
                    elif col_name != 'IPs':
                        columns_to_drop_due_to_concat.extend(existing_cols)
                    elif col_name != 'Host':
                        columns_to_drop_due_to_concat.extend(existing_cols)
                    else:
                        logger.info(f'Columns kept: {existing_cols}')
                if missing_cols:
                    logger.info(f'Columns not found in "{col_name}": {missing_cols}')
            if columns_to_drop_due_to_concat:
                # ! Drop columns that were used to create new columns
                df.drop(columns=columns_to_drop_due_to_concat, inplace=True)
                dropped_columns = available_columns_before_drop - set(df.columns)
                logger.info(f'Columns dropped from concat: {columns_to_drop_due_to_concat}')
                if dropped_columns:
                    logger.info("Columns dropped from concat:")
                    for column in dropped_columns:
                        logger.info(column)
                else:
                    logger.info("No columns were dropped.")
            logger.info('[red]########### columns_to_drop  from concat completed ###########[/red]')
            logger.info(f'Columns dropped.\nFinal DF: {df.shape[0]} rows {df.shape[1]} columns')
            logger.info(f'[green]Columns to be dropped started: {columns_to_be_dropped}[/green]')
            available_columns_before_drop2 = set(df.columns)
            cols_to_drop_from_func_def = [column for column in columns_to_be_dropped if column in df.columns]
            # ! Drop columns specified in the function definition
            if cols_to_drop_from_func_def:
                df.drop(columns=cols_to_drop_from_func_def, inplace=True)
                dropped_columns2 = available_columns_before_drop2 - set(df.columns)
                if dropped_columns2:
                    logger.info("Columns dropped:")
                    for column in dropped_columns2:
                        logger.info(column)
                logger.info(f'Columns dropped 2.\nFinal DF: {df.shape[0]} rows {df.shape[1]} columns')
                logger.info('[green]Columns to be dropped ended: [/green]')
            return df

        def apply_final_column_order_and_save(df: pd.DataFrame, final_column_order: List[str], output_file: str, available_columns_before_drop: set) -> None:
            final_order_logs = []
            final_order_logs.append('Final order started')
            final_order_logs.append('actual_final_order')
            actual_final_order = [col for col in final_column_order if col in df.columns]
            final_order_logs.append(','.join(actual_final_order))
            final_order_logs.append('specified_columns')
            specified_columns = set(actual_final_order)
            final_order_logs.append(','.join(specified_columns))
            final_order_logs.append('unspecified_columns')
            unspecified_columns = available_columns_before_drop - specified_columns
            final_order_logs.append(','.join(unspecified_columns))
            final_order_logs.append('unspecified_columns')
            unspecified_columns = unspecified_columns - set(columns_to_be_dropped)
            final_order_logs.append(','.join(unspecified_columns))
            final_order_logs.append('final_columns_list')

            final_columns_list = actual_final_order + [col for col in list(unspecified_columns) if col in df.columns]
            final_order_logs.append(','.join(final_columns_list))

            with open(logfile, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(final_order_logs)
                logger.info(f'Final order log file created: {logfile}')
            logger.info(f'Final columns list: {final_columns_list}')
            try:
                # ! Subset df with the final columns list
                df = df[final_columns_list]

                # ! Proceed with saving the file
                df.to_excel(output_file, index=False)
                logger.info(f'Excel file created: {output_file}')
            except KeyError as e:
                logger.error(f"Error in applying final column order: {e}")
                # ! Handle the error or perform additional logging as needed

        df: pd.DataFrame = pd.read_excel(input_file)
        # ! Section 1
        columns_to_drop_due_to_concat: List[str] = []
        available_columns_before_drop = set(df.columns)
        df_total_rows: int = len(df)
        df_total_columns: int = len(df.columns)
        new_columns: Dict[str, Dict[str, Any]] = {
            'Qs': {
                'cols': ['requestQuery_s', 'originalRequestUriWithArgs_s', 'requestUri_s'],
                'include_names': True
            },
            'IPs': {
                'cols': ['clientIp_s', 'clientIP_s'],
                'include_names': False
            },
            'Host': {
                'cols': ['host_s', 'hostname_s', 'originalHost_s'],
                'include_names': False
            },
            'messagesD': {
                'cols': ['Message', 'action_s'],
                'include_names': True
            },
            'allD': {
                'cols': ['details_message_s', 'details_data_s', 'details_file_s', 'details_line_s'],
                'include_names': True
            },
            'rulesD': {
                'cols': ['engine_s', 'ruleSetVersion_s', 'ruleGroup_s', 'ruleId_s', 'ruleSetType_s', 'policyId_s', 'policyScope_s', 'policyScopeName_s'],
                'include_names': True
            },
            'serverD': {
                'cols': ['backendSettingName_s', 'noOfConnectionRequests_d', 'serverRouted_s', 'httpVersion_s'],
                'include_names': True
            },
            'Method_Type': {
                'cols': ['httpMethod_s', 'contentType_s'],
                'include_names': False
            },
            'TimeGenerated_eventD': {
                'cols': [
                    'eventTimestamp_s', 'errorMessage_s', 'errorCount_d', 'Environment_s', 'Region_s', 'ActivityName_s', 'operationalResult_s', 'ActivityId_g', 'ScaleUnit_s',
                    'NamespaceName_s'
                ],
                'include_names': True
            },
        }
        logger.info('[red]########## Final Excel Creation Started #########[/red]')
        logger.info(f'DF loaded from file: {input_file}')
        logger.info(f'Total Columns: {df_total_columns} - Total Rows: {df_total_rows}')
        # ! Section 2
        drop_and_create_columns(df, new_columns)
        # ! Section 3
        apply_final_column_order_and_save(df, final_column_order, output_file, available_columns_before_drop)

    @staticmethod
    def read_and_filter_excel(file_path, columns=None) -> list[Any]:
        """
        Reads specified columns from an Excel file and returns unique non-blank values.
        """
        if columns is None:
            columns = ['', '']
        try:
            df = pd.read_excel(file_path, usecols=columns)
            unique_data = pd.unique(df[columns].values.ravel('K'))
            unique_data = [data for data in unique_data if pd.notna(data)]
            return unique_data
        except FileNotFoundError:
            print(f"Error: The file {file_path} was not found.")
        except EmptyDataError:
            print("Error: The file is empty.")
        except ValueError as e:
            print(f"Error: {e}. Check if the columns exist.")
        return []

    @staticmethod
    def copy_sheets_to_new_file(source_files, new_file):
        """
        Copies all sheets from given source Excel files to a new Excel file.
        """
        try:
            new_wb = Workbook()
            new_wb.remove(new_wb.active)  # type: ignore

            for file in source_files:
                wb = load_workbook(file)
                for sheet_name in wb.sheetnames:
                    source = wb[sheet_name]
                    target = new_wb.create_sheet(title=sheet_name)

                    for row in source:
                        for cell in row:
                            target[cell.coordinate].value = cell.value

            new_wb.save(new_file)
            print(f"New file {new_file} has been created with all copied sheets.")
        except FileNotFoundError as e:
            print(f"Error: {e}. One of the source files was not found.")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
